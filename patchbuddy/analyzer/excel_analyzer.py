import openpyxl

class ExcelAnalyzer:
    def __init__(self):
        pass

    def analyze(self, filepath):
        try:
            # We need data_only=False to see formulas
            wb = openpyxl.load_workbook(filepath, data_only=False, read_only=True)
            sheets = {}
            sheet_order = wb.sheetnames
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Column headers and order
                headers = []
                # ws[1] gives a tuple of cells in the first row
                # In read_only mode, we should be careful with ws[1] if it's empty
                try:
                    for cell in ws[1]:
                        val = str(cell.value) if cell.value is not None else ""
                        headers.append(val)
                except Exception:
                    pass # Handle empty sheets
                
                # Track formulas
                formulas = {}
                # In read_only mode, using iter_rows is much faster than ws.cell()
                max_r = min(ws.max_row or 0, 1000)
                max_c = min(ws.max_column or 0, 50)
                
                for row_idx, row in enumerate(ws.iter_rows(max_row=max_r, max_col=max_c), start=1):
                    for col_idx, cell in enumerate(row, start=1):
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            formulas[f"{openpyxl.utils.get_column_letter(col_idx)}{row_idx}"] = cell.value
                
                sheets[sheet_name] = {
                    "columns": headers,
                    "column_count": len(headers),
                    "row_count": ws.max_row,
                    "formulas": formulas,
                    "formula_count": len(formulas),
                    "merged_cells": [str(r) for r in ws.merged_cells.ranges]
                }
            
            return {
                "sheets": sheets,
                "sheet_order": sheet_order,
                "type": "excel"
            }
        except Exception as e:
            return {"error": str(e), "type": "excel"}
